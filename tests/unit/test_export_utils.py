# Version 1.0 - 29.07.2026 14:10:00 GMT
# Unit tests for services/database/export_utils.py
# Описание: Проверяет формирование столбцов tLib и pCloud в XLSX-экспорте:
#           наличие заголовков, корректность ссылок, пустые ячейки при отсутствии файла,
#           percent-кодирование имени файла, поведение при незаданном SITE_URL.

from __future__ import annotations

import sqlite3
from pathlib import Path

import pytest
from openpyxl import load_workbook

from config import DATABASE_TABLE_NAME
from services.database.tlib_table_spec import (
    build_create_table_sql,
    build_insert_sql,
    build_values,
)


# ---------------------------------------------------------------------------
# Хелперы
# ---------------------------------------------------------------------------


def _make_db(path: Path, rows: list[dict]) -> str:
    """Создаёт tlib.db с переданными строками, возвращает строковый путь."""
    conn = sqlite3.connect(str(path))
    conn.execute(build_create_table_sql(DATABASE_TABLE_NAME))
    for row in rows:
        conn.execute(build_insert_sql(DATABASE_TABLE_NAME), build_values(row))
    conn.commit()
    conn.close()
    return str(path)


def _export(db_path: str, xlsx_path: str) -> None:
    """Запускает экспорт через тестируемую функцию."""
    from services.database.export_utils import export_database_to_xlsx
    result = export_database_to_xlsx(db_path, xlsx_path)
    assert result is True, "export_database_to_xlsx вернула False — экспорт не выполнен"


def _read_xlsx(xlsx_path: str):
    """Возвращает (headers, data_rows) из первого листа XLSX."""
    wb = load_workbook(xlsx_path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    data = [list(r) for r in rows[1:]]
    return headers, data


def _col(headers: list, name: str) -> int:
    """Индекс столбца по имени (0-based)."""
    return headers.index(name)


# ---------------------------------------------------------------------------
# Тесты
# ---------------------------------------------------------------------------


class TestExtraColumnsPresent:
    """Проверяет, что tLib и pCloud — последние два заголовка."""

    def test_headers_end_with_tlib_pcloud(self, tmp_path, monkeypatch):
        import services.database.export_utils as m
        monkeypatch.setattr(m, "SITE_URL", "https://tlib.ru")
        monkeypatch.setattr(m, "PCLOUD_DATA_BASE_URL", "https://example.com/data/")
        db = _make_db(tmp_path / "tlib.db", [{"Шифр": 1, "ДопШифр": "TST", "РазмерАрхива": 100, "ТипФайла": "zip"}])
        xlsx = str(tmp_path / "out.xlsx")
        _export(db, xlsx)
        headers, _ = _read_xlsx(xlsx)
        assert headers[-2] == "tLib"
        assert headers[-1] == "pCloud"


class TestTlibUrl:
    """Проверяет формирование столбца tLib."""

    def _run(self, tmp_path, monkeypatch, rows, site_url="https://tlib.ru"):
        import services.database.export_utils as m
        monkeypatch.setattr(m, "SITE_URL", site_url)
        monkeypatch.setattr(m, "PCLOUD_DATA_BASE_URL", "https://example.com/data/")
        db = _make_db(tmp_path / "tlib.db", rows)
        xlsx = str(tmp_path / "out.xlsx")
        _export(db, xlsx)
        headers, data = _read_xlsx(xlsx)
        return headers, data

    def test_with_dopshifr(self, tmp_path, monkeypatch):
        headers, data = self._run(tmp_path, monkeypatch, [
            {"Шифр": 1, "ДопШифр": "TST", "РазмерАрхива": 0, "ТипФайла": ""},
        ])
        url = data[0][_col(headers, "tLib")]
        assert url == "https://tlib.ru/?1-TST"

    def test_without_dopshifr(self, tmp_path, monkeypatch):
        headers, data = self._run(tmp_path, monkeypatch, [
            {"Шифр": 42, "ДопШифр": "", "РазмерАрхива": 0, "ТипФайла": ""},
        ])
        url = data[0][_col(headers, "tLib")]
        assert url == "https://tlib.ru/?42"

    def test_empty_site_url_gives_empty_cell(self, tmp_path, monkeypatch):
        headers, data = self._run(tmp_path, monkeypatch, [
            {"Шифр": 1, "ДопШифр": "TST", "РазмерАрхива": 0, "ТипФайла": ""},
        ], site_url="")
        url = data[0][_col(headers, "tLib")]
        assert not url


class TestPCloudUrl:
    """Проверяет формирование столбца pCloud."""

    def _run(self, tmp_path, monkeypatch, rows, pcloud_base="https://example.com/data/"):
        import services.database.export_utils as m
        monkeypatch.setattr(m, "SITE_URL", "https://tlib.ru")
        monkeypatch.setattr(m, "PCLOUD_DATA_BASE_URL", pcloud_base)
        db = _make_db(tmp_path / "tlib.db", rows)
        xlsx = str(tmp_path / "out.xlsx")
        _export(db, xlsx)
        headers, data = _read_xlsx(xlsx)
        return headers, data

    def test_zip_with_dopshifr(self, tmp_path, monkeypatch):
        headers, data = self._run(tmp_path, monkeypatch, [
            {"Шифр": 1, "ДопШифр": "TST", "РазмерАрхива": 500, "ТипФайла": "zip"},
        ])
        url = data[0][_col(headers, "pCloud")]
        assert url == "https://example.com/data/00001-TST.zip"

    def test_pdf_without_dopshifr(self, tmp_path, monkeypatch):
        headers, data = self._run(tmp_path, monkeypatch, [
            {"Шифр": 3617, "ДопШифр": "", "РазмерАрхива": 1024, "ТипФайла": "pdf"},
        ])
        url = data[0][_col(headers, "pCloud")]
        assert url == "https://example.com/data/03617.pdf"

    def test_cyrillic_dopshifr_is_encoded(self, tmp_path, monkeypatch):
        headers, data = self._run(tmp_path, monkeypatch, [
            {"Шифр": 1, "ДопШифр": "ш", "РазмерАрхива": 100, "ТипФайла": "pdf"},
        ])
        url = data[0][_col(headers, "pCloud")]
        assert "%D1%88" in url or "%d1%88" in url.lower()
        assert "00001" in url

    def test_no_file_gives_empty_cell_zero_razmer(self, tmp_path, monkeypatch):
        headers, data = self._run(tmp_path, monkeypatch, [
            {"Шифр": 1, "ДопШифр": "TST", "РазмерАрхива": 0, "ТипФайла": "zip"},
        ])
        url = data[0][_col(headers, "pCloud")]
        assert not url

    def test_no_file_gives_empty_cell_no_tipfayla(self, tmp_path, monkeypatch):
        headers, data = self._run(tmp_path, monkeypatch, [
            {"Шифр": 1, "ДопШифр": "TST", "РазмерАрхива": 500, "ТипФайла": ""},
        ])
        url = data[0][_col(headers, "pCloud")]
        assert not url
