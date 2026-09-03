# Version 2.0 - 29.07.2026 14:00:00 GMT
# Export Utils - Экспорт базы данных в документы
# Описание: Модуль для экспорта базы данных SQLite в различные форматы документов.
#           export_database_to_xlsx() экспортирует первую таблицу БД в XLSX файл с форматированием,
#           автоподбором ширины колонок, очисткой недопустимых символов и экранированием формул.
#           В конце таблицы — два вычисляемых столбца: tLib (ссылка на страницу отчёта на сайте)
#           и pCloud (ссылка на файл отчёта в облачном зеркале). Обе — кликабельные гиперссылки.
#           Зависит от библиотеки openpyxl. Изолирован от критической логики обновления БД
#           для предотвращения блокировки системных операций при проблемах с openpyxl.
# 2.0: строки читаются через sqlite3.Row; добавлены столбцы tLib и pCloud с гиперссылками.

import re
import sqlite3
import urllib.parse
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from logging_config import app_logger
from config import PCLOUD_DATA_BASE_URL, SITE_URL
from services.seo.report_seo import build_canonical_query

# Имена вычисляемых столбцов-ссылок, добавляемых после колонок БД
_EXTRA_COLUMNS = ("tLib", "pCloud")

# Стиль гиперссылок в XLSX
_HYPERLINK_FONT = Font(color="0563C1", underline="single")

# Внимание: лимит Excel на гиперссылки — ~65 530 на лист.
# При двух столбцах ссылок лимит исчерпается при ~32 700 отчётах.
# Пока записей ~12 тысяч — запас двукратный; при росте пересмотреть стратегию.


def _build_tlib_url(row: sqlite3.Row) -> str:
    """
    Возвращает URL страницы отчёта вида {SITE_URL}/?{Шифр}-{ДопШифр}.
    Пустая строка, если SITE_URL не задан или Шифр некорректен.
    """
    if not SITE_URL:
        return ""
    try:
        canonical = build_canonical_query(dict(row))
        return f"{SITE_URL.rstrip('/')}/?" + urllib.parse.quote(canonical, safe="-")
    except Exception:
        return ""


def _build_pcloud_url(row: sqlite3.Row) -> str:
    """
    Возвращает URL файла отчёта в облачном зеркале.
    Пустая строка, если файл отсутствует (РазмерАрхива пуст/ноль или ТипФайла пуст).
    """
    razmer = row["РазмерАрхива"]
    tip = (row["ТипФайла"] or "").strip()
    if not razmer or not tip:
        return ""
    try:
        shifr = int(row["Шифр"])
    except (TypeError, ValueError):
        return ""
    dop = (row["ДопШифр"] or "").strip()
    filename = f"{shifr:05d}-{dop}.{tip}" if dop else f"{shifr:05d}.{tip}"
    encoded = urllib.parse.quote(filename, safe="")
    base = PCLOUD_DATA_BASE_URL.rstrip("/") + "/"
    return base + encoded


def export_database_to_xlsx(db_path: str, xlsx_path: str) -> bool:
    """
    Экспортирует первую таблицу из SQLite базы данных в XLSX файл.

    Процесс:
    1. Открывает БД и получает список таблиц
    2. Экспортирует первую таблицу (без столбца id) + два вычисляемых столбца tLib/pCloud
    3. Форматирует заголовки и автоподбирает ширину колонок
    4. Для столбцов tLib и pCloud расставляет кликабельные гиперссылки
    5. Сохраняет XLSX файл

    Args:
        db_path: Путь к SQLite базе данных
        xlsx_path: Путь для сохранения XLSX файла

    Returns:
        bool: True если экспорт успешен, False при ошибке
    """
    def sanitize_cell_value(value):
        """Очищает значение ячейки от недопустимых символов и экранирует формулы."""
        if value is None:
            return None
        if isinstance(value, str):
            cleaned = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F]', '', value)
            if cleaned and cleaned[0] in ('=', '+', '-', '@', '\t'):
                cleaned = "'" + cleaned
            return cleaned
        return value

    try:
        app_logger.info(f"Экспорт БД в XLSX: {db_path} → {xlsx_path}")

        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row

        # Получаем список таблиц
        cursor = conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%';"
        )
        tables = [row[0] for row in cursor.fetchall()]

        if not tables:
            app_logger.warning("База данных не содержит таблиц для экспорта")
            conn.close()
            return False

        table = tables[0]
        app_logger.info(f"Экспорт таблицы: {table}")

        # Получаем данные таблицы
        cursor = conn.execute(f'SELECT * FROM "{table}"')
        db_rows = cursor.fetchall()

        # Имена колонок БД (без 'id')
        all_columns = [description[0] for description in cursor.description]
        db_columns = [c for c in all_columns if c != "id"]

        # Итоговый список заголовков = колонки БД + два вычисляемых столбца
        columns = db_columns + list(_EXTRA_COLUMNS)

        # Строки: значения колонок БД + два вычисленных URL
        rows = []
        for db_row in db_rows:
            values = [db_row[c] for c in db_columns]
            values.append(_build_tlib_url(db_row))
            values.append(_build_pcloud_url(db_row))
            rows.append(values)

        conn.close()

        # Создаём XLSX
        wb = Workbook()
        wb.remove(wb.active)

        date_str = datetime.now().strftime("%Y.%m.%d")
        ws = wb.create_sheet(title=date_str)

        # Заголовки с форматированием
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center')
        for col_idx, column in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=sanitize_cell_value(column))
            cell.font = header_font
            cell.alignment = header_alignment

        # Индексы вычисляемых столбцов (1-based)
        tlib_col_idx = len(columns) - 1
        pcloud_col_idx = len(columns)

        # Данные + гиперссылки для двух последних столбцов
        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=sanitize_cell_value(value))
                if col_idx in (tlib_col_idx, pcloud_col_idx) and value:
                    cell.hyperlink = value
                    cell.font = _HYPERLINK_FONT

        # Автоподбор ширины колонок
        for col_idx, column in enumerate(columns, 1):
            max_length = len(str(column))
            for row in rows:
                cell_value = row[col_idx - 1]
                if cell_value:
                    max_length = max(max_length, min(len(str(cell_value)), 50))
            ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

        wb.save(xlsx_path)

        app_logger.info(f"XLSX экспорт завершён: {len(rows)} записей, {len(columns)} колонок")
        return True

    except Exception as e:
        app_logger.error(f"Ошибка экспорта БД в XLSX: {e}", exc_info=True)
        return False
